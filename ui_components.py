import customtkinter as ctk
from tkinter import ttk, messagebox
import threading
from datetime import datetime
import os
from openpyxl import Workbook
from market_apis import YandexMarketAPI, OzonAPI, WildberriesAPI, StockRow
import asyncio

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")


class MarketplacePanel(ctk.CTkFrame):
    """Панель для одного маркетплейса"""

    def __init__(self, master, title, api_client, api_client_ozon=None, api_client_wb=None, cooldown_seconds=0, **kwargs):
        super().__init__(master, **kwargs)
        self.title = title
        self.api_client = api_client
        self.api_client_ozon = api_client_ozon
        self.api_client_wb = api_client_wb
        self.cooldown_seconds = cooldown_seconds
        self.cooldown_active = False
        self.cooldown_timer = None

        self.stocks = []
        self.loading = False

        self.setup_ui()

    def setup_ui(self):
        self.title_label = ctk.CTkLabel(
            self, text=self.title,
            font=ctk.CTkFont(size=16, weight="bold")
        )
        self.title_label.pack(pady=5)

        self.tree_frame = ctk.CTkFrame(self)
        self.tree_frame.pack(fill="both", expand=True, padx=5, pady=5)

        columns = ("offer_id", "available", "reserved")
        self.tree = ttk.Treeview(self.tree_frame, columns=columns, show="headings", height=15)

        self.tree.heading("offer_id", text="Артикул", command=lambda: self.sort_by("offer_id"))
        self.tree.heading("available", text="Остаток", command=lambda: self.sort_by("available"))
        self.tree.heading("reserved", text="Резерв", command=lambda: self.sort_by("reserved"))

        self.tree.column("offer_id", width=140, anchor="center")
        self.tree.column("available", width=80, anchor="center")
        self.tree.column("reserved", width=70, anchor="center")

        scrollbar = ttk.Scrollbar(self.tree_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)

        self.tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        self.update_btn = ctk.CTkButton(
            self, text="Обновить остатки",
            command=self.update_stocks,
            height=35
        )
        self.update_btn.pack(pady=5, padx=20, fill="x")

        self.export_btn = ctk.CTkButton(
            self, text="📊 Экспорт в Excel",
            command=self.export_to_excel,
            height=35,
            fg_color="#2d6a4f",
            hover_color="#1b4d3e"
        )
        self.export_btn.pack(pady=5, padx=20, fill="x")

        self.status_label = ctk.CTkLabel(self, text="Готов", font=ctk.CTkFont(size=10))
        self.status_label.pack(pady=2)

        self.timer_label = ctk.CTkLabel(
            self,
            text="",
            font=ctk.CTkFont(size=10),
            text_color="orange",
            height=20
        )
        self.timer_label.pack(pady=2)

        if self.cooldown_seconds <= 0:
            self.timer_label.configure(text="")
            self.timer_label.pack_forget()
            self.timer_placeholder = ctk.CTkFrame(self, height=20, fg_color="transparent")
            self.timer_placeholder.pack(pady=2)
        else:
            self.timer_placeholder = None

    def start_cooldown(self):
        """Запускает таймер ожидания"""
        if self.cooldown_seconds <= 0:
            return

        self.cooldown_active = True
        self.update_btn.configure(state="disabled", text=f"Подождите {self.cooldown_seconds}с...")

        self._update_timer_display(self.cooldown_seconds)

    def _update_timer_display(self, remaining):
        """Обновление отображения таймера"""
        if remaining <= 0:
            self.cooldown_active = False
            self.update_btn.configure(state="normal", text="Обновить остатки")
            if hasattr(self, 'timer_label'):
                self.timer_label.configure(text="")
        else:
            if hasattr(self, 'timer_label'):
                self.timer_label.configure(text=f"⏰ Ожидание: {remaining} сек")
            self.cooldown_timer = self.after(1000, lambda: self._update_timer_display(remaining - 1))

    def cancel_cooldown(self):
        """Отменяет таймер (если нужно)"""
        if self.cooldown_timer:
            self.after_cancel(self.cooldown_timer)
            self.cooldown_timer = None
        self.cooldown_active = False
        self.update_btn.configure(state="normal", text="Обновить остатки")
        if hasattr(self, 'timer_label'):
            self.timer_label.configure(text="")

    def export_to_excel(self):
        """Экспорт данных в Excel файл"""
        if not self.stocks:
            messagebox.showwarning("Нет данных", "Сначала обновите остатки!")
            return

        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{self.title}_остатки_{timestamp}.xlsx"
        filepath = os.path.join(desktop, filename)

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = f"{self.title} Остатки"

            ws['A1'] = "Артикул"
            ws['B1'] = "Остаток"
            ws['C1'] = "Резерв"
            ws['D1'] = "Дата выгрузки"

            from openpyxl.styles import Font
            for cell in ws[1]:
                cell.font = Font(bold=True)

            for row, stock in enumerate(self.stocks, start=2):
                ws[f'A{row}'] = stock.offer_id
                ws[f'B{row}'] = stock.available
                ws[f'C{row}'] = stock.reserved
                ws[f'D{row}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(filepath)

            messagebox.showinfo(
                "Экспорт завершён",
                f"Файл сохранён:\n{filepath}\n\n"
                f"Количество товаров: {len(self.stocks)}"
            )

            self.status_label.configure(text=f"Экспортировано: {filename}")

        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось сохранить файл:\n{e}")
            self.status_label.configure(text=f"Ошибка экспорта: {str(e)[:30]}")

    def sort_by(self, column):
        """Сортировка таблицы"""
        reverse = False
        if hasattr(self, '_last_sort_column') and self._last_sort_column == column:
            reverse = True if self._last_sort_column_reverse is False else False
            self._last_sort_column_reverse = reverse
        else:
            self._last_sort_column = column
            self._last_sort_column_reverse = False

        if column == "available" or column == "reserved":
            self.stocks.sort(key=lambda x: getattr(x, column), reverse=reverse)
        else:
            self.stocks.sort(key=lambda x: self.parse_article(x.offer_id), reverse=reverse)

        self.refresh_table()

    def refresh_table(self):
        """Обновление отображения таблицы"""
        for item in self.tree.get_children():
            self.tree.delete(item)

        for stock in self.stocks:
            tags = ()
            if stock.available == 0:
                tags = ("zero",)
            elif stock.available <= 5:
                tags = ("low",)
            elif stock.available <= 10:
                tags = ("medium",)

            self.tree.insert("", "end", values=(
                stock.offer_id,
                stock.available,
                stock.reserved
            ), tags=tags)

        self.tree.tag_configure("zero", background="#8B0000")
        self.tree.tag_configure("low", background="#D2691E")
        self.tree.tag_configure("medium", background="#B8860B")

    def update_stocks(self):
        """Обновление остатков (в отдельном потоке)"""
        if self.loading:
            return

        if self.cooldown_active:
            messagebox.showwarning(
                "Подождите",
                f"Пожалуйста, подождите {self.cooldown_seconds} секунд между запросами к API.\n"
                f"Это ограничение самого маркетплейса."
            )
            return

        self.loading = True
        self.update_btn.configure(state="disabled", text="Загрузка...")
        self.status_label.configure(text="Загрузка данных...")

        thread = threading.Thread(target=self._do_update)
        thread.daemon = True
        thread.start()


    def parse_article(self, article):
        """Парсит артикул вида BAZ-XX-YY-ZZ и возвращает кортеж чисел для сортировки"""
        try:
            if article.startswith("BAZ-"):
                parts = article[4:].split("-")
                numbers = tuple(int(part) for part in parts if part.isdigit())
                return numbers
            else:
                return (article,)
        except (ValueError, AttributeError):
            return (article,)

    def _do_update(self):
        """Фоновое обновление"""
        try:

            if self.title == "ОЗОН" and self.api_client_ozon:
                loop = asyncio.new_event_loop()
                asyncio.set_event_loop(loop)
                stocks = loop.run_until_complete(self.api_client_ozon.get_stocks())
                loop.close()
            elif self.title == "WILDBERRIES" and self.api_client_wb:
                loop = asyncio.new_event_loop()
                asyncio.set_event_loop(loop)
                stocks = loop.run_until_complete(self.api_client_wb.get_stocks())
                loop.close()
            else:
                loop = asyncio.new_event_loop()
                asyncio.set_event_loop(loop)
                stocks = loop.run_until_complete(self.api_client.get_stocks())
                loop.close()

            if stocks is not None:
                stocks.sort(key=lambda x: self.parse_article(x.offer_id))
                self.stocks = stocks
                self.after(0, self.refresh_table)
                self.after(0, lambda: self.status_label.configure(text=f"Обновлено: {len(stocks)} товаров"))

                if self.title == "WILDBERRIES":
                    self.after(0, self.start_cooldown)
            else:
                self.after(0, lambda: self.status_label.configure(text="Ошибка: слишком много запросов"))

        except Exception as e:
            self.after(0, lambda: self.status_label.configure(text=f"Ошибка: {str(e)[:50]}"))

        finally:
            self.loading = False
            if not self.cooldown_active:
                self.after(0, lambda: self.update_btn.configure(state="normal", text="Обновить остатки"))